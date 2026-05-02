#!/usr/bin/env python3
"""
Coastal Key Enterprise — AI Agent Fleet Registry
Generates Excel report from all agent definition files.

Run: python3 generate_fleet_registry.py
Output: CK_Enterprise_AI_Fleet_Registry.xlsx (400 agents, 17 columns, 2 sheets)
"""

import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

def extract_agents_from_js(filepath):
    with open(filepath, 'r') as f:
        content = f.read()
    agents = []
    pattern = r'\{[^{}]*?id:\s*[\'"]([^"\']+)[\'"][^{}]*?\}'
    for match in re.finditer(pattern, content, re.DOTALL):
        block = match.group(0)
        agent = {}
        for field in ['id', 'name', 'role', 'description', 'division', 'tier', 'status', 'squad']:
            m = re.search(rf"{field}:\s*['\"]([ ^'\"]*)['\"]" , block)
            if m:
                agent[field] = m.group(1)
        for arr_field in ['triggers', 'outputs', 'kpis']:
            arr_m = re.search(rf"{arr_field}:\s*\[([^\]]*)\]", block)
            if arr_m:
                agent[arr_field] = ', '.join(re.findall(r"'([^']*)'" , arr_m.group(1)))
        sev_m = re.search(r"severity:\s*['\"]([^'\"]*)['\"]" , block)
        if sev_m:
            agent['severity'] = sev_m.group(1)
        if agent.get('id'):
            agents.append(agent)
    return agents

def extract_campaign_agents(filepath):
    with open(filepath, 'r') as f:
        config = json.load(f)
    return [{
        'id': a['id'], 'name': a['name'], 'role': 'Retell AI Voice Sales Agent',
        'description': 'TH Sentinel outbound sales agent. Engages homeowners, qualifies leads, transfers to Tracey Hunter.',
        'division': 'TH-SEN', 'tier': 'retell', 'status': a.get('status', 'active'),
        'squad': 'TH Sentinel Campaign', 'triggers': 'Outbound call, inbound callback',
        'outputs': 'Lead qualification, call transfer, DNC flag',
        'kpis': 'Connection rate, qualification rate, transfer rate',
    } for a in config.get('agents', {}).get('agent_roster', [])]

DIVISION_META = {
    'EXC': {'full_name': 'Executive', 'supervisor': 'CEO / Board of Directors', 'color': '6366F1'},
    'SEN': {'full_name': 'Sentinel Sales', 'supervisor': 'VP of Sales (EXC-002)', 'color': 'EF4444'},
    'OPS': {'full_name': 'Operations', 'supervisor': 'COO (EXC-001)', 'color': 'F59E0B'},
    'INT': {'full_name': 'Intelligence', 'supervisor': 'CIO (EXC-003)', 'color': '10B981'},
    'MKT': {'full_name': 'Marketing', 'supervisor': 'CMO (EXC-017)', 'color': '8B5CF6'},
    'FIN': {'full_name': 'Finance', 'supervisor': 'CFO (EXC-004)', 'color': '06B6D4'},
    'VEN': {'full_name': 'Vendor Management', 'supervisor': 'VP Vendor Relations (EXC-006)', 'color': 'F97316'},
    'TEC': {'full_name': 'Technology', 'supervisor': 'CTO (EXC-012)', 'color': '64748B'},
    'WEB': {'full_name': 'Web Development', 'supervisor': 'CTO (EXC-012)', 'color': '0EA5E9'},
    'IO': {'full_name': 'Intelligence Officers', 'supervisor': 'CIO (EXC-003)', 'color': '059669'},
    'EMAIL': {'full_name': 'Email AI Operations', 'supervisor': 'CMO (EXC-017)', 'color': 'D946EF'},
    'TH-SEN': {'full_name': 'TH Sentinel Campaign', 'supervisor': 'Campaign Director (SEN-001)', 'color': 'DC2626'},
}

def assign_gender(agent):
    if agent.get('division') == 'TH-SEN': return 'F'
    num_match = re.search(r'(\d+)', agent.get('id', '0'))
    return 'M' if num_match and int(num_match.group(1)) % 2 == 1 else 'F'

def build_excel(all_agents, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Agent Fleet Registry'
    headers = ['Agent ID', 'Name', 'M/F', 'Title', 'Division', 'Division Code', 'Supervisor', 'Job Description', 'Platform', 'Tier', 'Live Status', 'Squad/Unit', 'Primary Triggers', 'Key Outputs', 'KPIs', 'Severity', 'Enterprise Notes']
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'), top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
    sc = {'active': 'DCFCE7', 'standby': 'FEF9C3', 'training': 'DBEAFE', 'maintenance': 'FEE2E2'}
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    for ri, agent in enumerate(all_agents, 2):
        dc = agent.get('division', 'N/A'); meta = DIVISION_META.get(dc, {'full_name': dc, 'supervisor': 'N/A', 'color': '94A3B8'})
        plat = 'Retell AI' if dc == 'TH-SEN' else ('Claude API (Specialized)' if dc in ('IO', 'EMAIL') else 'Claude API')
        g = assign_gender(agent); s = agent.get('status', 'active').capitalize()
        notes = []; 
        if agent.get('tier') == 'advanced': notes.append('Advanced tier (Opus)')
        if agent.get('severity') in ('critical', 'high'): notes.append(f"{agent['severity'].upper()} severity")
        if s.lower() == 'standby': notes.append('Standby')
        if dc == 'TH-SEN': notes.append('Retell AI platform')
        vals = [agent.get('id',''), agent.get('name',''), g, agent.get('role',''), meta['full_name'], dc, meta['supervisor'], agent.get('description',''), plat, agent.get('tier','standard').capitalize(), s, agent.get('squad','—'), agent.get('triggers','—'), agent.get('outputs','—'), agent.get('kpis','—'), (agent.get('severity','—') or '—').capitalize(), '; '.join(notes) or '—']
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val); c.font = Font(name='Calibri', size=10); c.alignment = Alignment(vertical='center', wrap_text=True); c.border = tb
            if col == 11 and s.lower() in sc: c.fill = PatternFill(start_color=sc[s.lower()], end_color=sc[s.lower()], fill_type='solid'); c.font = Font(name='Calibri', size=10, bold=True)
            if col == 1: c.fill = PatternFill(start_color=meta.get('color','94A3B8'), end_color=meta.get('color','94A3B8'), fill_type='solid'); c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    for i, w in enumerate([16,22,5,35,20,10,40,60,18,12,12,18,35,35,30,12,40], 1): ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_path)
    return len(all_agents)

if __name__ == '__main__':
    all_agents = []
    for f in ['agents-exc.js','agents-sen.js','agents-ops.js','agents-int.js','agents-mkt.js','agents-fin.js','agents-ven.js','agents-tec.js','agents-web.js']:
        fp = os.path.join(BASE, 'ck-api-gateway/src/agents', f)
        if os.path.exists(fp): all_agents.extend(extract_agents_from_js(fp))
    for f, div in [('ck-api-gateway/src/routes/intelligence-officers.js', 'IO'), ('ck-api-gateway/src/routes/email-agents.js', 'EMAIL')]:
        fp = os.path.join(BASE, f)
        if os.path.exists(fp):
            agents = extract_agents_from_js(fp)
            for a in agents: a['division'] = div
            all_agents.extend(agents)
    cp = os.path.join(BASE, 'th-sentinel-campaign/campaign-config.json')
    if os.path.exists(cp): all_agents.extend(extract_campaign_agents(cp))
    print(f'Total agents: {len(all_agents)}')
    out = os.path.join(BASE, 'CK_Enterprise_AI_Fleet_Registry.xlsx')
    build_excel(all_agents, out)
    print(f'Generated: {out}')